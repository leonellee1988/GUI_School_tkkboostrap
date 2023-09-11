import tkinter
from tkinter import *
import ttkbootstrap
from ttkbootstrap.dialogs import Messagebox
import openpyxl
import datetime
import pandas
import numpy

#Main window configuration:
window = ttkbootstrap.Window(themename='minty')
window.title('Creación de asignatura')
window.geometry('400x325')
window.resizable()

#My functions:
def label_config(text_label, font_label, x_label, y_label):
    """
    This function creates ttkbootstrap labels considering the text,
    font and the position of the label as arguments.
    
    >>>lable_config('Hello, I am a label', ('Helvetica', 28), 10, 10)
    """
    ttkbootstrap.Label(window, text=text_label, font=font_label).place(x=x_label, y=y_label)

def load_data():
    """
    This function lets the user to load information about a subject into an Excel file,
    setting a subject's code automatically.
    
    No arguments need for this function.
    """
    file = openpyxl.load_workbook('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
    sheet = file['materia']
    
    first_value = sheet.cell(column=1, row=sheet.max_row).value
    counter = 0
    if type(first_value) == type(''):
        counter = 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=1)
    else:
        counter = first_value + 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=counter)
    
    data_name = entry_name.get()
    data_mode = combobox_mode.get()
    data_credit = combobox_credit.get()
    data_cost = entry_cost.get()
    data_career = combobox_code_career.get()
    date = datetime.datetime.now()
    data_date = date.strftime('%d/%m/%Y %H:%M:%S')
    
    if data_name=='' or data_mode=='' or data_credit=='' or data_cost=='' or data_career=='':
        Messagebox.show_error('Falta información por ingresar', title='Error!')
    else:
        sheet.cell(column=2, row=sheet.max_row, value='MAT0000' + str(counter))
        sheet.cell(column=3, row=sheet.max_row, value=data_name)
        sheet.cell(column=4, row=sheet.max_row, value=data_mode)
        sheet.cell(column=5, row=sheet.max_row, value=data_credit)
        sheet.cell(column=6, row=sheet.max_row, value=data_cost)
        sheet.cell(column=7, row=sheet.max_row, value=data_date)
        sheet.cell(column=8, row=sheet.max_row, value=data_career)
        entry_name.delete(0, END)
        combobox_mode.delete(0, END)
        combobox_credit.delete(0, END)
        entry_cost.delete(0, END)
        combobox_code_career.delete(0, END)
        file.save('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
        Messagebox.show_info('Información cargada correctamente', title='Felicidades')

def career_check():
    """
    This functions helps the user to know what is the name of the career asociated with a code.
    
    No arguments taken.
    """
    career_name_nparray = (career_sheet[career_sheet['CODIGO_CARRERA']==combobox_code_career.get()]['NOMBRE_CARRERA']).values
    career_degree_nparray = (career_sheet[career_sheet['CODIGO_CARRERA']==combobox_code_career.get()]['GRADO_CARRERA']).values
    career_name_str = numpy.array_str(career_name_nparray)
    career_degree_str = numpy.array_str(career_degree_nparray)
    career_name_str_final = career_name_str[2:len(career_name_str)-2]
    career_degree_str_final = career_degree_str[2:len(career_degree_str)-2]
    if career_name_str_final != '':
        Messagebox.show_info('El código corresponde a: ' + career_name_str_final + ' ' + career_degree_str_final, title='Información carrera')
    else:
        Messagebox.show_error('Ingrese el código de la carrera.', title='Error!')
        
#-----------------------------------------------------------------------------------------------------------

#Main frame

main_frame = tkinter.LabelFrame(window, text='Creación de asignatura')
main_frame.place(x=10, y=10, width=375, height=300)

#Name field:
label_config('Materia', ('Helvetica', 11), 30, 50)
entry_name = ttkbootstrap.Entry(window)
entry_name.place(x=120, y=50)

#Mode field:
label_config('Modalidad', ('Helvetica', 11), 30, 90)
mode = ['PRESENCIAL', 'VIRTUAL', 'HIBRIDO']
combobox_mode = ttkbootstrap.Combobox(window,values=mode)
combobox_mode.place(x=120, y=90)

#Credits field:
label_config('Créditos', ('Helvetica', 11), 30, 130)
credit = [1, 2, 3, 4, 5]
combobox_credit = ttkbootstrap.Combobox(window, values=credit)
combobox_credit.place(x=120, y=130)

#Cost field:
label_config('Costo', ('Helvetica', 11), 30, 170)
entry_cost = ttkbootstrap.Entry(window)
entry_cost.place(x=120, y=170)

#Career code:
label_config('Código carrera', ('Helvetica', 11), 30, 210)
career_sheet = pandas.read_excel('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx', sheet_name='carrera')
code_career = career_sheet['CODIGO_CARRERA']
code_career_list = list(code_career)
combobox_code_career = ttkbootstrap.Combobox(window, values=code_career_list)
combobox_code_career.place(x=160, y=210)

#Load button:
load_button = ttkbootstrap.Button(text='Subir', command=load_data).place(x=110, y=260)

#Career button:
career_button = ttkbootstrap.Button(text='Carrera', command=career_check).place(x=210, y=260)

window.mainloop()