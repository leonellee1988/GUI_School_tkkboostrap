import tkinter
from tkinter import *
import ttkbootstrap
from ttkbootstrap.dialogs import Messagebox
import openpyxl
import datetime

#Main window configuration:
window = ttkbootstrap.Window(themename='minty')
window.title('Creación de facultad')
window.geometry('350x300')
window.resizable()

#My functions:
def label_config(text_label, font_label, x_label, y_label):
    """
    This function creates ttkbootstrap labels considering the text,
    font and the position of the label as arguments.
    
    >>>lable_config('Hello, I am a label', ('Helvetica', 28), 10, 10)
    """
    ttkbootstrap.Label(window, text=text_label, font=font_label).place(x=x_label, y=y_label)

def load_data():
    """
    This function lets the user to load information about a faculty into an Excel file,
    setting a faculty's code automatically.
    
    No arguments need for this function.
    """
    file = openpyxl.load_workbook('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
    sheet = file['facultad']
    
    first_value = sheet.cell(column=1, row=sheet.max_row).value
    counter = 0
    if type(first_value) == type(''):
        counter = 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=1)
    else:
        counter = first_value + 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=counter)
    
    data_name = entry_name.get()
    data_dean = entry_dean.get()
    data_phone = entry_phone.get()
    data_email = entry_email.get()
    date = datetime.datetime.now()
    data_date = date.strftime('%d/%m/%Y %H:%M:%S')
    if data_name=='' or data_dean=='' or data_phone=='' or data_email=='':
        Messagebox.show_error('Falta información por ingresar', title='Error!')
    elif len(data_phone) != 9:
        Messagebox.show_error('Ingrese el teléfono en formato 0000-0000', title='Error!')
    elif '@' not in data_email:
        Messagebox.show_error('Ingrese el correo electrónico en formato micorreo@mail.com', title='Error!')
    else:
        sheet.cell(column=2, row=sheet.max_row, value='FAC0000' + str(counter))
        sheet.cell(column=3, row=sheet.max_row, value=data_name)
        sheet.cell(column=4, row=sheet.max_row, value=data_dean)
        sheet.cell(column=5, row=sheet.max_row, value=data_phone)
        sheet.cell(column=6, row=sheet.max_row, value=data_email)
        sheet.cell(column=7, row=sheet.max_row, value=data_date)
        entry_name.delete(0, END)
        entry_dean.delete(0, END)
        entry_phone.delete(0, END)
        entry_email.delete(0, END)
        file.save('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
        Messagebox.show_info('Información cargada correctamente', title='Felicidades')
        
#-----------------------------------------------------------------------------------------------------------

#Main frame

main_frame = tkinter.LabelFrame(window, text='Creación de facultad')
main_frame.place(x=10, y=10, width=325, height=275)

#Name field:
label_config('Facultad', ('Helvetica', 11), 30, 50)
entry_name = ttkbootstrap.Entry(window)
entry_name.place(x=100, y=50)

#Dean field:
label_config('Decano', ('Helvetica', 11), 30, 90)
entry_dean = ttkbootstrap.Entry(window)
entry_dean.place(x=100, y=90)

#Phone field:
label_config('Teléfono', ('Helvetica', 11), 30, 130)
entry_phone = ttkbootstrap.Entry(window)
entry_phone.place(x=100, y=130)

#Email field:
label_config('Email', ('Helvetica', 11), 30, 170)
entry_email = ttkbootstrap.Entry(window)
entry_email.place(x=100, y=170)

#Load button:
load_button = ttkbootstrap.Button(window, text='Subir', command=load_data).place(x=150, y=230)

window.mainloop()