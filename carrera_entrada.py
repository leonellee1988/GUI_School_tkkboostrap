import tkinter
from tkinter import *
import ttkbootstrap
from ttkbootstrap.dialogs import Messagebox
import openpyxl
import datetime
import pandas
import numpy

#Main window configuration:
window = ttkbootstrap.Window(themename='minty')
window.title('Creación de carrera')
window.geometry('390x285')
window.resizable()

#My functions:
def label_config(text_label, font_label, x_label, y_label):
    """
    This function creates ttkbootstrap labels considering the text,
    font and the position of the label as arguments.
    
    >>>lable_config('Hello, I am a label', ('Helvetica', 28), 10, 10)
    """
    ttkbootstrap.Label(window, text=text_label, font=font_label).place(x=x_label, y=y_label)

def load_data():
    """
    This function lets the user to load information about a career into an Excel file,
    setting a career's code automatically.
    
    No arguments need for this function.
    """
    file = openpyxl.load_workbook('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
    sheet = file['carrera']
    
    first_value = sheet.cell(column=1, row=sheet.max_row).value
    counter = 0
    if type(first_value) == type(''):
        counter = 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=1)
    else:
        counter = first_value + 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=counter)
    
    data_name = entry_name.get()
    data_degree = combobox_degree.get()
    data_term = combobox_term.get()
    data_faculty = combobox_code_faculty.get()
    date = datetime.datetime.now()
    data_date = date.strftime('%d/%m/%Y %H:%M:%S')
    
    if data_name=='' or data_degree=='' or data_term=='' or data_faculty=='':
        Messagebox.show_error('Falta información por ingresar', title='Error!')
    else:
        sheet.cell(column=2, row=sheet.max_row, value='CAR0000' + str(counter))
        sheet.cell(column=3, row=sheet.max_row, value=data_name)
        sheet.cell(column=4, row=sheet.max_row, value=data_degree)
        sheet.cell(column=5, row=sheet.max_row, value=data_term)
        sheet.cell(column=6, row=sheet.max_row, value=data_date)
        sheet.cell(column=7, row=sheet.max_row, value=data_faculty)
        entry_name.delete(0, END)
        combobox_degree.delete(0, END)
        combobox_term.delete(0, END)
        combobox_code_faculty.delete(0, END)
        file.save('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
        Messagebox.show_info('Información cargada correctamente', title='Felicidades')

def faculty_check():
    """
    This functions helps the user to know what is the name of the faculty asociated with a code.
    
    No arguments taken.
    """
    faculty_name_nparray = (faculty_sheet[faculty_sheet['CODIGO_FACULTAD']==combobox_code_faculty.get()]['NOMBRE_FACULTAD']).values
    faculty_name_str = numpy.array_str(faculty_name_nparray)
    faculty_name_str_final = faculty_name_str[2:len(faculty_name_str)-2]
    if faculty_name_str_final != '':
        Messagebox.show_info('El código corresponde a: ' + faculty_name_str_final, title='Información facultad')
    else:
        Messagebox.show_error('Ingrese el código de facultad.', title='Error!')
        
#-----------------------------------------------------------------------------------------------------------

#Main frame

main_frame = tkinter.LabelFrame(window, text='Creación de carrera')
main_frame.place(x=10, y=10, width=365, height=260)

#Name field:
label_config('Carrera', ('Helvetica', 11), 30, 50)
entry_name = ttkbootstrap.Entry(window)
entry_name.place(x=100, y=50)

#Degree field:
label_config('Grado', ('Helvetica', 11), 30, 90)
degree = ['TÉCNICO', 'LICENCIATURA', 'MAESTRÍA', 'DOCTORADO']
combobox_degree = ttkbootstrap.Combobox(window,values=degree)
combobox_degree.place(x=100, y=90)

#Term field:
label_config('Duración', ('Helvetica', 11), 30, 130)
term = [1, 2, 3, 4, 5]
combobox_term = ttkbootstrap.Combobox(window, values=term)
combobox_term.place(x=100, y=130)

#Faculty code:
label_config('Código facultad', ('Helvetica', 11), 30, 170)
faculty_sheet = pandas.read_excel('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx', sheet_name='facultad')
code_faculty = faculty_sheet['CODIGO_FACULTAD']
code_faculty_list = list(code_faculty)
combobox_code_faculty = ttkbootstrap.Combobox(window, values=code_faculty_list)
combobox_code_faculty.place(x=145, y=170)

#Load button:
load_button = ttkbootstrap.Button(text='Subir', command=load_data).place(x=110, y=220)

#Faculty button:
faculty_button = ttkbootstrap.Button(text='Facultad', command=faculty_check).place(x=210, y=220)

window.mainloop()