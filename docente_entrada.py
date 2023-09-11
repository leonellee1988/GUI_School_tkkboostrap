import tkinter
from tkinter import *
import ttkbootstrap
from ttkbootstrap.dialogs import Messagebox
import openpyxl
import datetime
import pandas
import numpy

#Main window configuration:
window = ttkbootstrap.Window(themename='minty')
window.title('Creación de docente')
window.geometry('375x410')
window.resizable()

#My functions:
def label_config(text_label, font_label, x_label, y_label):
    """
    This function creates ttkbootstrap labels considering the text,
    font and the position of the label as arguments.
    
    >>>lable_config('Hello, I am a label', ('Helvetica', 28), 10, 10)
    """
    ttkbootstrap.Label(window, text=text_label, font=font_label).place(x=x_label, y=y_label)

def load_data():
    """
    This function lets the user to load information about a teacher into an Excel file,
    setting a teacher's code automatically.
    
    No arguments need for this function.
    """
    file = openpyxl.load_workbook('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
    sheet = file['docente']
    
    first_value = sheet.cell(column=1, row=sheet.max_row).value
    counter = 0
    if type(first_value) == type(''):
        counter = 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=1)
    else:
        counter = first_value + 1
        sheet.cell(column=1, row=sheet.max_row + 1, value=counter)
    
    data_name = entry_name.get()
    data_last_name = entry_last_name.get()
    data_gender = combobox_gender.get()
    data_birthday = entry_birthday.get()
    data_phone = entry_phone.get()
    data_email = entry_email.get()
    data_faculty = combobox_code_faculty.get()
    date = datetime.datetime.now()
    data_date = date.strftime('%d/%m/%Y %H:%M:%S')
    
    if data_name=='' or data_last_name=='' or data_gender=='' or data_birthday=='' or data_phone=='' or data_email=='' or data_faculty=='':
        Messagebox.show_error('Falta información por ingresar', title='Error!')
    elif '/' not in data_birthday:
        Messagebox.show_error('Ingrese la fecha de nacimiento en formato 00/00/0000', title='Error!')
    elif len(data_phone) != 9:
        Messagebox.show_error('Ingrese el teléfono en formato 0000-0000', title='Error!')
    elif '@' not in data_email:
        Messagebox.show_error('Ingrese el correo electrónico en formato micorreo@mail.com', title='Error!')
    else:
        sheet.cell(column=2, row=sheet.max_row, value='DOC0000' + str(counter))
        sheet.cell(column=3, row=sheet.max_row, value=data_name)
        sheet.cell(column=4, row=sheet.max_row, value=data_last_name)
        sheet.cell(column=5, row=sheet.max_row, value=data_gender)
        sheet.cell(column=6, row=sheet.max_row, value=data_birthday)
        sheet.cell(column=7, row=sheet.max_row, value=data_phone)
        sheet.cell(column=8, row=sheet.max_row, value=data_email)
        sheet.cell(column=9, row=sheet.max_row, value=data_date)
        sheet.cell(column=10, row=sheet.max_row, value=data_faculty)
        entry_name.delete(0, END)
        entry_last_name.delete(0, END)
        combobox_gender.delete(0, END)
        entry_birthday.delete(0, END)
        entry_phone.delete(0, END)
        entry_email.delete(0, END)
        combobox_code_faculty.delete(0, END)
        file.save('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx')
        Messagebox.show_info('Información cargada correctamente', title='Felicidades')

def faculty_check():
    """
    This functions helps the user to know what is the name of the faculty asociated with a code.
    
    No arguments taken.
    """
    faculty_name_nparray = (faculty_sheet[faculty_sheet['CODIGO_FACULTAD']==combobox_code_faculty.get()]['NOMBRE_FACULTAD']).values
    faculty_name_str = numpy.array_str(faculty_name_nparray)
    faculty_name_str_final = faculty_name_str[2:len(faculty_name_str)-2]
    if faculty_name_str_final != '':
        Messagebox.show_info('El código corresponde a: ' + faculty_name_str_final, title='Información facultad')
    else:
        Messagebox.show_error('Ingrese el código de facultad.', title='Error!')
        
#-----------------------------------------------------------------------------------------------------------

#Main frame

main_frame = tkinter.LabelFrame(window, text='Creación de docente')
main_frame.place(x=10, y=10, width=355, height=385)

#Name field:
label_config('Nombre', ('Helvetica', 11), 30, 50)
entry_name = ttkbootstrap.Entry(window)
entry_name.place(x=100, y=50)

#Last name field:
label_config('Apellido', ('Helvetica', 11), 30, 90)
entry_last_name = ttkbootstrap.Entry(window)
entry_last_name.place(x=100, y=90)

#Gender field:
label_config('Genero', ('Helvetica', 11), 30, 130)
gender = ['FEMENINO', 'MASCULINO']
combobox_gender = ttkbootstrap.Combobox(window, values=gender)
combobox_gender.place(x=100, y=130)

#Birthday field:
label_config('Fecha Nac.', ('Helvetica', 11), 30, 170)
entry_birthday = ttkbootstrap.Entry(window)
entry_birthday.place(x=125, y=170)

#Phone field:
label_config('Teléfono', ('Helvetica', 11), 30, 210)
entry_phone = ttkbootstrap.Entry(window)
entry_phone.place(x=100, y=210)

#Email field:
label_config('Email', ('Helvetica', 11), 30, 250)
entry_email = ttkbootstrap.Entry(window)
entry_email.place(x=100, y=250)

#Faculty code:
label_config('Código facultad', ('Helvetica', 11), 30, 290)
faculty_sheet = pandas.read_excel('/home/bruce/Escritorio/Python/GUI/BD_SCHOOL_2023.xlsx', sheet_name='facultad')
code_faculty = faculty_sheet['CODIGO_FACULTAD']
code_faculty_list = list(code_faculty)
combobox_code_faculty = ttkbootstrap.Combobox(window, values=code_faculty_list)
combobox_code_faculty.place(x=140, y=290)

#Load button:
load_button = ttkbootstrap.Button(text='Subir', command=load_data).place(x=110, y=350)

#Faculty button:
faculty_button = ttkbootstrap.Button(text='Facultad', command=faculty_check).place(x=210, y=350)

window.mainloop()